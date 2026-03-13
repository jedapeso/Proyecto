import json
import os
import tempfile
import time
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError
from datetime import datetime
from threading import Lock

import pandas as pd
from dotenv import load_dotenv
from flask import jsonify, render_template, request, send_file
from google import genai
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from src import engine
from src.extensions import redis_client

from . import hospitalizacion_bp

load_dotenv()

api_key = os.getenv("GOOGLE_API_KEY")
gemini_model_name = os.getenv("GEMINI_MODEL", "gemini-1.5-flash")
_gemini_client = None
_gemini_executor = ThreadPoolExecutor(max_workers=2)

CENSO_CACHE_TIMEOUT_SECONDS = 300
GEMINI_CACHE_TIMEOUT_SECONDS = 1800
GEMINI_TIMEOUT_SECONDS = 10
RANGO_EDAD_LABELS = ["0-17 anos", "18-39 anos", "40-59 anos", "60+ anos"]

_local_cache = {}
_cache_locks = {}
_cache_locks_guard = Lock()

HEADER_FONT = Font(bold=True)
ALTERNATE_FILL = PatternFill(fill_type="solid", fgColor="F2F2F2")
DAY_FILL_LOW = PatternFill(fill_type="solid", fgColor="2ECC71")
DAY_FILL_MEDIUM = PatternFill(fill_type="solid", fgColor="FFF176")
DAY_FILL_HIGH = PatternFill(fill_type="solid", fgColor="F5B041")
DAY_FILL_CRITICAL = PatternFill(fill_type="solid", fgColor="E74C3C")


def get_gemini_client():
    """Inicializa el cliente Gemini solo cuando realmente se necesita."""
    global _gemini_client

    if _gemini_client is not None:
        return _gemini_client

    if not api_key:
        return None

    try:
        _gemini_client = genai.Client(api_key=api_key)
    except Exception:
        _gemini_client = None

    return _gemini_client


def get_json_payload():
    """Evita fallos cuando el body JSON viene vacio o invalido."""
    return request.get_json(silent=True) or {}


def build_servicio_filter(column_name, servicio):
    """Devuelve un WHERE seguro para servicio sin interpolar valores del request."""
    if servicio == "TS":
        return f"WHERE {column_name} <> 'TS'", {}
    return f"WHERE {column_name} = :servicio", {"servicio": servicio}


def normalize_empresa(empresa):
    value = (empresa or "").strip()
    return value or "0"


def get_cache_lock(cache_key):
    with _cache_locks_guard:
        if cache_key not in _cache_locks:
            _cache_locks[cache_key] = Lock()
        return _cache_locks[cache_key]


def cache_get(cache_key):
    if redis_client is not None:
        cached = redis_client.get(cache_key)
        if cached:
            try:
                return json.loads(cached).get("value")
            except json.JSONDecodeError:
                return None

    fallback_value = _local_cache.get(cache_key)
    if fallback_value is None:
        return None
    if fallback_value["expires_at"] <= time.time():
        _local_cache.pop(cache_key, None)
        return None
    return fallback_value["value"]


def cache_set(cache_key, value, timeout_seconds):
    payload = json.dumps({"value": value}, ensure_ascii=True, default=str)
    if redis_client is not None:
        redis_client.setex(cache_key, timeout_seconds, payload)
        return

    _local_cache[cache_key] = {
        "value": value,
        "expires_at": time.time() + timeout_seconds,
    }


def build_censo_cache_key(servicio, empresa):
    return f"hospitalizacion:censo:{servicio}:{normalize_empresa(empresa)}"


def build_gemini_cache_key(paciente_id, texto):
    base_token = (paciente_id or "").strip()
    if not base_token:
        base_token = str(abs(hash(texto.strip())))
    return f"hospitalizacion:gemini:{base_token}"


def empty_analisis_result():
    return {
        "asegurador": {"labels": [], "valores": []},
        "rango_edad": {"labels": RANGO_EDAD_LABELS, "valores": [0, 0, 0, 0]},
        "diagnosticos": {"labels": [], "valores": []},
        "cruzado": {"aseguradores": [], "rangos": RANGO_EDAD_LABELS, "datos": {}},
    }


def execute_censo_procedure(conn, servicio, empresa):
    conn.execute(
        text("EXECUTE PROCEDURE SP_Censod(:servicio, :empresa)"),
        {"servicio": servicio, "empresa": normalize_empresa(empresa)},
    )


def fetch_censo_summary_dataframe(conn, servicio):
    where_clause, query_params = build_servicio_filter("C1.UBI", servicio)
    query = f"""
        SELECT UNIQUE
            CASE
                WHEN C1.UBI = 'H1' THEN 1
                WHEN C1.UBI = 'H2' THEN 2
                WHEN C1.UBI = 'H3' THEN 3
                WHEN C1.UBI = 'UA' THEN 4
                ELSE 6
            END AS ORDEN_UBI,
            C1.UBI,
            C1.OCUPADAS,
            C1.DISPONIBLES,
            C1.NO_HAB_SERV,
            C1.POR_OCU_SER,
            C1.POR_OCU_SER_TOT,
            C1.PROM_ESTANCIA,
            COALESCE(C2.CNT_ALTA, 0) AS CNT_ALTA
        FROM CENSO1 C1
        LEFT JOIN (
            SELECT UBI, COUNT(*) AS CNT_ALTA
            FROM CENSO1
            WHERE DIAS_ESTANCIA > 10
              AND ASEGURADOR IS NOT NULL
              AND TRIM(ASEGURADOR) <> ''
            GROUP BY UBI
        ) C2 ON C2.UBI = C1.UBI
        {where_clause}
        ORDER BY ORDEN_UBI
    """
    return pd.read_sql(text(query), conn, params=query_params)


def fetch_censo_analysis_dataframe(conn, servicio):
    where_clause, query_params = build_servicio_filter("UBI", servicio)
    query = f"""
        SELECT ASEGURADOR, EDAD, DIAGNOSTICO, CIE10
        FROM CENSO1
        {where_clause}
        AND ASEGURADOR IS NOT NULL
        AND TRIM(ASEGURADOR) <> ''
    """
    return pd.read_sql(text(query), conn, params=query_params)


def fetch_censo_report_dataframes(servicio, empresa):
    with engine.begin() as conn:
        execute_censo_procedure(conn, servicio, empresa)
        df_censo = pd.read_sql(
            text(
                """
                SELECT SERVICIO,
                       HABITACION,
                       HISTORIA,
                       IDENTIFICACION,
                       DIAS_ESTANCIA,
                       NOMBRE_COMPLETO,
                       SEXO,
                       GENERO,
                       ASEGURADOR,
                       EDAD,
                       CIE10,
                       DIAGNOSTICO
                FROM CENSO1
                WHERE UBI <> 'TS'
                ORDER BY SERVICIO, HABITACION
                """
            ),
            conn,
        )
        df_resumen = pd.read_sql(
            text(
                """
                SELECT UNIQUE
                    CASE
                        WHEN UBI = 'H1' THEN 1
                        WHEN UBI = 'H2' THEN 2
                        WHEN UBI = 'H3' THEN 3
                        WHEN UBI = 'UA' THEN 4
                        WHEN UBI = 'TS' THEN 5
                        ELSE 6
                    END AS NO,
                    SERVICIO,
                    OCUPADAS,
                    DISPONIBLES,
                    FUERA_SERVICIO,
                    NO_HAB_SERV,
                    POR_OCU_SER,
                    POR_OCU_SER_TOT,
                    PROM_ESTANCIA
                FROM CENSO1
                ORDER BY NO
                """
            ),
            conn,
        )

    df_censo.columns = df_censo.columns.str.upper()
    df_resumen.columns = df_resumen.columns.str.upper()

    if "DIAS_ESTANCIA" in df_censo.columns:
        df_censo["DIAS_ESTANCIA"] = pd.to_numeric(df_censo["DIAS_ESTANCIA"], errors="coerce")

    for column_name in ["PROM_ESTANCIA", "POR_OCU_SER", "POR_OCU_SER_TOT"]:
        if column_name in df_resumen.columns:
            df_resumen[column_name] = pd.to_numeric(df_resumen[column_name], errors="coerce")

    if not df_resumen.empty:
        df_resumen.rename(
            columns={
                "SERVICIO": "Ubicacion",
                "OCUPADAS": "Camas Ocupadas",
                "DISPONIBLES": "Camas Disponibles",
                "FUERA_SERVICIO": "Camas Fuera de Servicio",
                "NO_HAB_SERV": "Nr Camas por Servicio",
                "POR_OCU_SER": "% Ocupacion Servicio",
                "POR_OCU_SER_TOT": "% Ocupacion General",
                "PROM_ESTANCIA": "Promedio Estancia",
            },
            inplace=True,
        )
        if "Promedio Estancia" in df_resumen.columns:
            df_resumen["Promedio Estancia"] = pd.to_numeric(df_resumen["Promedio Estancia"], errors="coerce")

    return df_censo, df_resumen


def serialize_censo_summary(df_summary):
    if df_summary.empty:
        return {
            "labels": [],
            "ocupadas": [],
            "disponibles": [],
            "porcentaje": [],
            "porcentaje_gral": [],
            "total": [],
            "estancia": [],
            "estancia_alta": [],
        }

    df_summary.columns = df_summary.columns.str.lower()
    df_summary = df_summary.fillna(0)

    for column_name in ["ocupadas", "disponibles", "no_hab_serv", "cnt_alta"]:
        df_summary[column_name] = pd.to_numeric(df_summary[column_name], errors="coerce").fillna(0).astype(int)

    for column_name in ["por_ocu_ser", "por_ocu_ser_tot", "prom_estancia"]:
        df_summary[column_name] = pd.to_numeric(df_summary[column_name], errors="coerce").fillna(0)

    df_summary["ubi"] = df_summary["ubi"].astype(str).str.strip()

    return {
        "labels": df_summary["ubi"].tolist(),
        "ocupadas": df_summary["ocupadas"].tolist(),
        "disponibles": df_summary["disponibles"].tolist(),
        "porcentaje": df_summary["por_ocu_ser"].tolist(),
        "porcentaje_gral": df_summary["por_ocu_ser_tot"].tolist(),
        "total": df_summary["no_hab_serv"].tolist(),
        "estancia": df_summary["prom_estancia"].tolist(),
        "estancia_alta": df_summary["cnt_alta"].tolist(),
    }


def serialize_censo_analysis(df_analysis):
    if df_analysis.empty:
        return empty_analisis_result()

    df_analysis.columns = df_analysis.columns.str.upper()
    df_analysis = df_analysis.fillna({"DIAGNOSTICO": "SIN DIAGNOSTICO"})
    df_analysis["EDAD"] = pd.to_numeric(df_analysis["EDAD"], errors="coerce").fillna(0).astype(int)
    df_analysis["RANGO_EDAD"] = pd.cut(
        df_analysis["EDAD"],
        bins=[-1, 17, 39, 59, 150],
        labels=RANGO_EDAD_LABELS,
    )

    top_aseguradores = df_analysis["ASEGURADOR"].value_counts().head(10)
    rango_counts = df_analysis["RANGO_EDAD"].value_counts(sort=False).reindex(RANGO_EDAD_LABELS, fill_value=0)
    diagnosticos_counts = df_analysis["DIAGNOSTICO"].value_counts().head(10)

    aseguradores_cruzados = df_analysis["ASEGURADOR"].value_counts().head(5).index.tolist()
    crosstab = pd.crosstab(df_analysis["ASEGURADOR"], df_analysis["RANGO_EDAD"]).reindex(
        index=aseguradores_cruzados,
        columns=RANGO_EDAD_LABELS,
        fill_value=0,
    )

    return {
        "asegurador": {
            "labels": top_aseguradores.index.tolist(),
            "valores": [int(value) for value in top_aseguradores.tolist()],
        },
        "rango_edad": {
            "labels": RANGO_EDAD_LABELS,
            "valores": [int(value) for value in rango_counts.tolist()],
        },
        "diagnosticos": {
            "labels": diagnosticos_counts.index.tolist(),
            "valores": [int(value) for value in diagnosticos_counts.tolist()],
        },
        "cruzado": {
            "aseguradores": aseguradores_cruzados,
            "rangos": RANGO_EDAD_LABELS,
            "datos": {
                asegurador: {rango: int(crosstab.loc[asegurador, rango]) for rango in RANGO_EDAD_LABELS}
                for asegurador in aseguradores_cruzados
            },
        },
    }


def get_censo_dashboard_payload(servicio, empresa):
    cache_key = build_censo_cache_key(servicio, empresa)
    cached_payload = cache_get(cache_key)
    if cached_payload is not None:
        return cached_payload

    with get_cache_lock(cache_key):
        cached_payload = cache_get(cache_key)
        if cached_payload is not None:
            return cached_payload

        with engine.begin() as conn:
            execute_censo_procedure(conn, servicio, empresa)
            df_summary = fetch_censo_summary_dataframe(conn, servicio)
            df_analysis = fetch_censo_analysis_dataframe(conn, servicio)

        payload = {
            "grafico": serialize_censo_summary(df_summary),
            "analisis": serialize_censo_analysis(df_analysis),
        }
        cache_set(cache_key, payload, CENSO_CACHE_TIMEOUT_SECONDS)
        return payload


def pick_fill_for_days(value):
    if value is None or pd.isna(value):
        return None

    numeric_value = float(value)
    if 1 <= numeric_value <= 5:
        return DAY_FILL_LOW
    if 6 <= numeric_value <= 9:
        return DAY_FILL_MEDIUM
    if 10 <= numeric_value <= 14:
        return DAY_FILL_HIGH
    if numeric_value >= 15:
        return DAY_FILL_CRITICAL
    return None


def set_openpyxl_column_widths(worksheet, dataframe, max_width=60):
    if dataframe.empty:
        for column_index, column_name in enumerate(dataframe.columns, start=1):
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(len(str(column_name)) + 2, max_width)
        return

    text_lengths = dataframe.astype(str).apply(lambda column: column.str.len().max())
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        max_len = max(int(text_lengths.iloc[column_index - 1]), len(str(column_name)))
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max_len + 2, max_width)


def build_write_only_row(worksheet, values, *, header=False, base_fill=None, percentage_indexes=None, day_indexes=None):
    percentage_indexes = percentage_indexes or set()
    day_indexes = day_indexes or set()
    row = []

    for column_index, value in enumerate(values):
        cell = WriteOnlyCell(worksheet, value=value)
        if header:
            cell.font = HEADER_FONT
        if base_fill is not None:
            cell.fill = base_fill
        if column_index in percentage_indexes and value not in (None, ""):
            cell.number_format = "0.00"
        if column_index in day_indexes:
            fill = pick_fill_for_days(value)
            if fill is not None:
                cell.fill = fill
        row.append(cell)

    return row


def append_dataframe_sheet(
    worksheet,
    dataframe,
    *,
    alternate_by_service=False,
    percentage_headers=None,
    heatmap_headers=None,
):
    percentage_headers = set(percentage_headers or [])
    heatmap_headers = set(heatmap_headers or [])
    worksheet.append(build_write_only_row(worksheet, list(dataframe.columns), header=True))

    if dataframe.empty:
        return

    service_groups = None
    if alternate_by_service and "SERVICIO" in dataframe.columns:
        service_groups = dataframe["SERVICIO"].ne(dataframe["SERVICIO"].shift()).cumsum().tolist()

    percentage_indexes = {
        index for index, header in enumerate(dataframe.columns) if header in percentage_headers
    }
    day_indexes = {index for index, header in enumerate(dataframe.columns) if header in heatmap_headers}

    normalized_frame = dataframe.where(pd.notnull(dataframe), None)
    for row_index, row_values in enumerate(normalized_frame.itertuples(index=False, name=None)):
        base_fill = None
        if service_groups is not None and service_groups[row_index] % 2 == 0:
            base_fill = ALTERNATE_FILL
        worksheet.append(
            build_write_only_row(
                worksheet,
                row_values,
                base_fill=base_fill,
                percentage_indexes=percentage_indexes,
                day_indexes=day_indexes,
            )
        )


def build_censo_excel_file(servicio, empresa):
    df_censo, df_resumen = fetch_censo_report_dataframes(servicio, empresa)
    output = tempfile.SpooledTemporaryFile(max_size=8_000_000, mode="w+b")
    workbook = Workbook(write_only=True)

    ws_detalle = workbook.create_sheet(title="Censo Hospitalario")
    append_dataframe_sheet(
        ws_detalle,
        df_censo,
        alternate_by_service=True,
        heatmap_headers={"DIAS_ESTANCIA"},
    )
    set_openpyxl_column_widths(ws_detalle, df_censo)

    ws_resumen = workbook.create_sheet(title="Resumen DATO1")
    append_dataframe_sheet(
        ws_resumen,
        df_resumen,
        percentage_headers={"% Ocupacion Servicio", "% Ocupacion General"},
        heatmap_headers={"Promedio Estancia"},
    )
    set_openpyxl_column_widths(ws_resumen, df_resumen)

    workbook.save(output)
    output.seek(0)
    return output


@hospitalizacion_bp.route("/", methods=["GET"])
def dashboard():
    return render_template("hospitalizacion/dashboard.html")


@hospitalizacion_bp.route("/censo", methods=["GET"])
def censo():
    with engine.begin() as conn:
        conn.execute(text("EXECUTE PROCEDURE SP_UbicaCensod()"))
        servicios = conn.execute(text("SELECT UBICOD, UBINOM FROM UBICAH1")).fetchall()

    return render_template("hospitalizacion/censo.html", servicios=servicios)


@hospitalizacion_bp.route("/empresas_por_servicio", methods=["POST"])
def get_empresas():
    data = get_json_payload()
    servicio = data.get("serv")

    if not servicio:
        return jsonify({"error": "Falta el parametro SERV"}), 400

    with engine.begin() as conn:
        conn.execute(text("EXECUTE PROCEDURE SP_ListadoEmpresasPAd(:serv)"), {"serv": servicio})
        empresas = conn.execute(text("SELECT EMPNIT, NITRAZ FROM NIT1")).fetchall()

    return jsonify([{"id": row[0], "nombre": row[1]} for row in empresas])


@hospitalizacion_bp.route("/servicios_por_empresa", methods=["POST"])
def servicios_por_empresa():
    data = get_json_payload()
    empresa = normalize_empresa(data.get("empresa"))

    if empresa == "0":
        with engine.begin() as conn:
            conn.execute(text("EXECUTE PROCEDURE SP_UbicaCensod()"))
            servicios = conn.execute(text("SELECT UBICOD, UBINOM FROM UBICAH1")).fetchall()
        return jsonify([{"ubi": row[0], "nombre": row[1]} for row in servicios])

    try:
        with engine.begin() as conn:
            conn.execute(
                text("EXECUTE PROCEDURE SP_Censod(:servicio, :empresa)"),
                {"servicio": "TS", "empresa": empresa},
            )
            rows = conn.execute(
                text(
                    """
                    SELECT UBI, COUNT(*) AS CNT
                    FROM CENSO1
                    WHERE UBI <> 'TS'
                      AND ASEGURADOR IS NOT NULL
                      AND TRIM(ASEGURADOR) <> ''
                    GROUP BY UBI
                    HAVING COUNT(*) > 0
                    ORDER BY
                        CASE
                            WHEN UBI='H1' THEN 1
                            WHEN UBI='H2' THEN 2
                            WHEN UBI='H3' THEN 3
                            WHEN UBI='UA' THEN 4
                            ELSE 9
                        END
                    """
                )
            ).fetchall()

        nombres = {
            "H1": "Hospitalizacion Piso 1",
            "H2": "Hospitalizacion Piso 2",
            "H3": "Hospitalizacion Piso 3",
            "UA": "UCI",
        }
        servicios_activos = [
            {"ubi": row[0].strip(), "nombre": nombres.get(row[0].strip(), row[0].strip())}
            for row in rows
        ]

        if len(servicios_activos) > 1:
            servicios_activos.insert(0, {"ubi": "TS", "nombre": "Todos los servicios"})

        return jsonify(servicios_activos)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@hospitalizacion_bp.route("/reporte_censo", methods=["POST"])
def reporte_censo():
    servicio = (request.form.get("servicio") or "").strip()
    empresa = normalize_empresa(request.form.get("empresa"))

    if not servicio:
        return "Debe seleccionar servicio", 400

    output = build_censo_excel_file(servicio, empresa)
    fecha_actual = datetime.now().strftime("%Y-%m-%d")
    nombre_archivo = f"CENSO_HOSP_{servicio}_{empresa}_{fecha_actual}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=nombre_archivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@hospitalizacion_bp.route("/datos_dashboard_censo", methods=["POST"])
def datos_dashboard_censo():
    data = get_json_payload()
    servicio = data.get("servicio")
    empresa = data.get("empresa")

    if not servicio:
        return jsonify({"error": "Debe seleccionar una ubicacion"}), 400

    try:
        return jsonify(get_censo_dashboard_payload(servicio, empresa))
    except Exception as exc:
        return jsonify({"error": f"Error al obtener datos: {exc}"}), 500


@hospitalizacion_bp.route("/datos_censo_grafico", methods=["POST"])
def datos_censo_grafico():
    data = get_json_payload()
    servicio = data.get("servicio")
    empresa = data.get("empresa")

    if not servicio:
        return jsonify({"error": "Debe seleccionar una ubicacion"}), 400

    try:
        payload = get_censo_dashboard_payload(servicio, empresa)
        return jsonify(payload["grafico"])
    except Exception as exc:
        return jsonify({"error": f"Error al obtener datos: {exc}"}), 500


@hospitalizacion_bp.route("/datos_analisis_adicionales", methods=["POST"])
def datos_analisis_adicionales():
    data = get_json_payload()
    servicio = data.get("servicio")
    empresa = data.get("empresa")

    if not servicio:
        return jsonify({"error": "Debe seleccionar una ubicacion"}), 400

    try:
        payload = get_censo_dashboard_payload(servicio, empresa)
        return jsonify(payload["analisis"])
    except Exception as exc:
        return jsonify({"error": f"Error al obtener analisis: {exc}"}), 500


@hospitalizacion_bp.route("/uci", methods=["GET"])
def tablero_uci():
    try:
        with engine.begin() as conn:
            result = conn.execute(text("SELECT camcod, camdes FROM ucicam ORDER BY camdes"))
            camas = result.fetchall()
        return render_template("hospitalizacion/tablero_uci.html", camas=camas)
    except Exception as exc:
        return f"Error cargando las camas: {exc}", 500


@hospitalizacion_bp.route("/uci/datos", methods=["POST"])
def obtener_datos_uci():
    try:
        data = get_json_payload()
        camcod = data.get("camcod")

        if not camcod:
            return jsonify({"error": "Falta parametro camcod"}), 400

        with engine.begin() as conn:
            conn.execute(text("EXECUTE PROCEDURE SP_Escalas_Ucid(:camcod)"), {"camcod": camcod})

            result_pacientes = conn.execute(
                text(
                    "SELECT ESCHIS,ESCNUM,ESCIDE,ESCPAC,ESCHAB,ESCDIA,ESCEPI,ESCAPA,ESAINT,"
                    "ESARGB,ESCBRA,ESBINT,ESBRGB,ESCRIE,ESRINT,ESRRGB,DIAEST "
                    "FROM ESCALAS1 ORDER BY ESCHAB"
                )
            )
            pacientes = result_pacientes.mappings().all()

            total_pacientes = conn.execute(text("SELECT TOTAL_PACIENTES FROM TOTPAC1")).scalar()
            promedio_apache = conn.execute(text("SELECT PROMEDIO_APACHE FROM PROAPA1")).scalar()
            riesgo_lpp = conn.execute(text("SELECT NO_PACI,TRIM(ESCALAPOR) ESCALAPOR FROM ESCUPP1")).mappings().all()
            riesgo_caida = conn.execute(text("SELECT NO_PACI,TRIM(ESCALAPOR) ESCALAPOR FROM RIECA1")).mappings().all()
            promedio_estancia = conn.execute(text("SELECT PROM FROM PROMES1")).scalar()

        response_data = {
            "pacientes_en_cama": [dict(row) for row in pacientes],
            "indicadores": {
                "total_pacientes": total_pacientes,
                "promedio_apache": promedio_apache,
                "riesgo_lpp": [dict(row) for row in riesgo_lpp],
                "riesgo_caida": [dict(row) for row in riesgo_caida],
                "promedio_estancia": str(promedio_estancia) if promedio_estancia is not None else "0",
            },
        }
        return jsonify(response_data)
    except Exception as exc:
        return jsonify({"error": f"Error al ejecutar SP_Escalas_Ucid: {exc}"}), 500


def convertir_a_lenguaje_natural(texto, paciente_id=None):
    if not texto or texto.strip().lower() in {"none", "null", ""}:
        return "Sin informacion disponible"

    cache_key = build_gemini_cache_key(paciente_id, texto)
    cached = cache_get(cache_key)
    if cached:
        return cached

    client = get_gemini_client()
    if client is None:
        return "La funcion de resumen IA no esta disponible."

    prompt = f"""
    Eres un asistente medico especializado en cuidados intensivos.
    Analiza el siguiente texto y devuelve:
    1. Riesgos del paciente.
    2. Necesidades prioritarias.
    3. Plan de cuidado concreto y entendible para enfermeria.
    Responde de forma clara y concisa.
    --- TEXTO ---
    {texto}
    """

    def _generate_summary():
        response = client.models.generate_content(
            model=gemini_model_name,
            contents=prompt,
        )
        if response and getattr(response, "text", None):
            return response.text.strip()
        return "Sin informacion generada"

    future = _gemini_executor.submit(_generate_summary)
    try:
        resumen = future.result(timeout=GEMINI_TIMEOUT_SECONDS)
    except FuturesTimeoutError:
        future.cancel()
        resumen = "El analisis IA excedio 10 segundos y se devolvio una respuesta parcial."
    except Exception:
        resumen = "El analisis IA no esta disponible en este momento."

    cache_set(cache_key, resumen, GEMINI_CACHE_TIMEOUT_SECONDS)
    return resumen


@hospitalizacion_bp.route("/uci/riesgos-necesidades-detalle", methods=["POST"])
def obtener_riesgos_necesidades():
    try:
        data = get_json_payload()
        escide = data.get("escide")

        if not escide:
            return jsonify({"error": "Falta parametro escide"}), 400

        with engine.begin() as conn:
            paciente = conn.execute(
                text(
                    """
                    SELECT ESCIDE, ESCPAC, ESCHAB, ESCDIA, ESCEPI
                    FROM ESCALAS1
                    WHERE ESCIDE = :escide
                    """
                ),
                {"escide": escide},
            ).fetchone()

            if not paciente:
                return jsonify({"resumen": "No se encontraron datos del paciente."}), 404

            escide_db, nombre, habitacion, diagnostico, escepi = paciente

            conn.execute(
                text("EXECUTE PROCEDURE SP_Riesgos_Necesidadesd(:ESCEPI)"),
                {"ESCEPI": escepi},
            )
            filas = conn.execute(text("SELECT LISTAS FROM RESUMEN1")).fetchall()

        if not filas:
            return jsonify(
                {
                    "escide": escide_db,
                    "nombre": nombre,
                    "habitacion": habitacion,
                    "diagnostico": diagnostico,
                    "resumen_ia": "No se encontraron riesgos o necesidades para este paciente.",
                }
            )

        riesgos_texto = " ".join(fila[0] for fila in filas if fila[0])
        resumen_ia = convertir_a_lenguaje_natural(riesgos_texto, paciente_id=str(escide_db))

        return jsonify(
            {
                "escide": escide_db,
                "nombre": nombre,
                "habitacion": habitacion,
                "diagnostico": diagnostico,
                "resumen_ia": resumen_ia,
            }
        )
    except Exception as exc:
        return jsonify({"error": f"Error al obtener riesgos: {exc}"}), 500
